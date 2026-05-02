# webservice.py
# Dot2Dot HTTP API 服务

from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
import os
import random
from datetime import datetime
import re
import time

# 强制切换到当前文件所在目录（解决路径问题）
os.chdir(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
CORS(app)

# ============================================================
# 静态页面路由
# ============================================================

@app.route('/')
@app.route('/login.html')
def index():
    with open('login.html', 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/homepage.html')
def homepage():
    with open('homepage.html', 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/postdetail.html')
def postdetail():
    with open('postdetail.html', 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/dot2dot.html')
def dot2dot():
    with open('dot2dot.html', 'r', encoding='utf-8') as f:
        return f.read()

# ============================================================
# 配置
# ============================================================

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'database.xlsx')
USERS_SHEET = 'sheet1'
MESSAGES_SHEET = 'sheet2'

WEIGHT_A = 3
WEIGHT_B = 5
WEIGHT_C = 10


# ============================================================
# 辅助函数
# ============================================================

def ensure_excel_exists():
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = USERS_SHEET
    ws1.append(['user_id', 'nickname', 'login_lastdate', 'login_times', 
                'aipost_lastdate', 'aipost_times', 'user_weight', 'user_tag', 
                'count_A', 'count_B', 'count_C'])
    ws2 = wb.create_sheet(MESSAGES_SHEET)
    ws2.append(['message_id', 'user_id', 'message_subject', 'message_content', 
                'message_tag', 'ai_processed'])
    wb.save(EXCEL_PATH)

def get_users():
    ensure_excel_exists()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[USERS_SHEET]
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        users.append({
            'user_id': row[0],
            'nickname': row[1],
            'login_lastdate': str(row[2]) if row[2] else '',
            'login_times': row[3] if row[3] else 0,
            'aipost_lastdate': str(row[4]) if row[4] else '',
            'aipost_times': row[5] if row[5] else 0,
            'user_weight': row[6] if row[6] is not None else 10,
            'user_tag': row[7] if row[7] else '',
            'count_A': row[8] if row[8] else 0,
            'count_B': row[9] if row[9] else 0,
            'count_C': row[10] if row[10] else 0
        })
    return users

def get_messages():
    ensure_excel_exists()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[MESSAGES_SHEET]
    messages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        messages.append({
            'message_id': row[0],
            'user_id': row[1],
            'message_subject': row[2] if row[2] else '',
            'message_content': row[3] if row[3] else '',
            'message_tag': row[4] if row[4] else '',
            'ai_processed': row[5] if row[5] else 0
        })
    return messages

def save_user(user):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[USERS_SHEET]
    target_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == user['user_id']:
            target_row = row_idx
            break
    if target_row:
        ws.cell(row=target_row, column=1, value=user['user_id'])
        ws.cell(row=target_row, column=2, value=user['nickname'])
        ws.cell(row=target_row, column=3, value=user['login_lastdate'])
        ws.cell(row=target_row, column=4, value=user['login_times'])
        ws.cell(row=target_row, column=5, value=user.get('aipost_lastdate', ''))
        ws.cell(row=target_row, column=6, value=user.get('aipost_times', 0))
        ws.cell(row=target_row, column=7, value=user['user_weight'])
        ws.cell(row=target_row, column=8, value=user.get('user_tag', ''))
        ws.cell(row=target_row, column=9, value=user.get('count_A', 0))
        ws.cell(row=target_row, column=10, value=user.get('count_B', 0))
        ws.cell(row=target_row, column=11, value=user.get('count_C', 0))
    else:
        ws.append([
            user['user_id'], user['nickname'], user['login_lastdate'], user['login_times'],
            user.get('aipost_lastdate', ''), user.get('aipost_times', 0),
            user['user_weight'], user.get('user_tag', ''), user.get('count_A', 0),
            user.get('count_B', 0), user.get('count_C', 0)
        ])
    wb.save(EXCEL_PATH)

def parse_message_tag(message_tag):
    if not message_tag:
        return {'tag': '', 'scores': [0, 0, 0, 0]}
    match = re.match(r'^(.*?)\s*\[(\d),(\d),(\d),(\d)\]$', message_tag)
    if match:
        return {
            'tag': match.group(1),
            'scores': [int(match.group(2)), int(match.group(3)), 
                       int(match.group(4)), int(match.group(5))]
        }
    return {'tag': message_tag, 'scores': [0, 0, 0, 0]}

def get_broadcast_quality(message_tag):
    parsed = parse_message_tag(message_tag)
    return sum(parsed['scores'])

def get_user_weight_by_id(user_id, users):
    for u in users:
        if u['user_id'] == user_id:
            return u['user_weight']
    return 5

def calculate_display_score(message, users):
    quality = get_broadcast_quality(message['message_tag'])
    user_weight = get_user_weight_by_id(message['user_id'], users)
    return quality * (user_weight / 10)

def is_broadcast_visible(message):
    parsed = parse_message_tag(message['message_tag'])
    return 0 not in parsed['scores']

def weighted_random_sample(items, weights, k):
    if len(items) <= k:
        return items.copy()
    remaining_items = items.copy()
    remaining_weights = weights.copy()
    total_weight = sum(remaining_weights)
    result = []
    for _ in range(k):
        if total_weight <= 0 or not remaining_items:
            break
        r = random.random() * total_weight
        acc = 0
        for i, w in enumerate(remaining_weights):
            acc += w
            if r <= acc:
                result.append(remaining_items.pop(i))
                total_weight -= remaining_weights.pop(i)
                break
    return result

def contains_chinese(text):
    """判断是否包含汉字"""
    return any('\u4e00' <= ch <= '\u9fff' for ch in text)

def extract_keywords(text):
    """
    从文本中提取关键词
    - 中文：按单字拆分
    - 英文/数字：按单词拆分（保持完整）
    """
    if not text:
        return []
    
    # 常见动词（作为重要特征保留）
    verbs = ['约', '找', '求', '学', '读', '看', '听', '玩', '做', '招募', '体验', '思考']
    
    keywords = []
    remaining = text
    
    # 检查是否以动词开头
    for v in verbs:
        if text.startswith(v):
            keywords.append(v)
            remaining = text[len(v):]
            break
    
    # 分词：区分中英文
    i = 0
    length = len(remaining)
    while i < length:
        ch = remaining[i]
        if contains_chinese(ch):
            # 中文：单字作为一个关键词
            keywords.append(ch)
            i += 1
        elif ch.isalnum() or ch == '_':
            # 英文/数字：累积一个完整单词
            start = i
            while i < length and (remaining[i].isalnum() or remaining[i] == '_'):
                i += 1
            word = remaining[start:i]
            if word:
                keywords.append(word)
        else:
            # 其他字符（空格、标点、@、.等）：跳过
            i += 1
    
    return list(set(keywords))

def is_related(tag1, tag2):
    """判断两个 tag 是否相关（有共同关键词）"""
    if not tag1 or not tag2:
        return False
    kw1 = extract_keywords(tag1)
    kw2 = extract_keywords(tag2)
    return bool(set(kw1) & set(kw2))


# ============================================================
# API 路由
# ============================================================

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    email = data.get('email', '').strip().lower()
    if not email or '@' not in email:
        return jsonify({'error': '无效的邮箱'}), 400
    nickname = email.split('@')[0]
    today = datetime.now().strftime('%Y-%m-%d')
    users = get_users()
    user = None
    for u in users:
        if u['user_id'] == email:
            user = u
            break
    if user:
        user['login_lastdate'] = today
        user['login_times'] = user.get('login_times', 0) + 1
    else:
        user = {
            'user_id': email,
            'nickname': nickname,
            'login_lastdate': today,
            'login_times': 1,
            'aipost_lastdate': '',
            'aipost_times': 0,
            'user_weight': 10,
            'user_tag': '',
            'count_A': 0,
            'count_B': 0,
            'count_C': 0
        }
    save_user(user)
    return jsonify({'status': 'ok', 'nickname': nickname, 'user_id': email})

@app.route('/api/users')
def api_users():
    return jsonify(get_users())

@app.route('/api/messages')
def api_messages():
    return jsonify(get_messages())

@app.route('/api/platform')
def api_platform():
    """平台广场：加权随机 10 条可见广播"""
    users = get_users()
    messages = get_messages()
    messages.sort(key=lambda x: x['message_id'], reverse=True)
    visible = [m for m in messages if is_broadcast_visible(m)]
    if not visible:
        return jsonify([])
    scores = [calculate_display_score(m, users) for m in visible]
    selected = weighted_random_sample(visible, scores, 10)
    user_dict = {u['user_id']: u for u in users}
    for item in selected:
        user = user_dict.get(item['user_id'], {})
        item['nickname'] = user.get('nickname', item['user_id'].split('@')[0])
    return jsonify(selected)

@app.route('/api/user/<user_id>/recommend')
def api_user_recommend(user_id):
    """用户视图推荐：加权随机 3 条该用户的可见广播"""
    users = get_users()
    messages = get_messages()
    user_messages = [m for m in messages if m['user_id'] == user_id and is_broadcast_visible(m)]
    if not user_messages:
        return jsonify([])
    scores = [calculate_display_score(m, users) for m in user_messages]
    selected = weighted_random_sample(user_messages, scores, 3)
    user_dict = {u['user_id']: u for u in users}
    for item in selected:
        user = user_dict.get(item['user_id'], {})
        item['nickname'] = user.get('nickname', item['user_id'].split('@')[0])
    return jsonify(selected)

@app.route('/api/user/<user_id>')
def api_user(user_id):
    users = get_users()
    for u in users:
        if u['user_id'] == user_id:
            return jsonify(u)
    return jsonify({'error': 'User not found'}), 404

@app.route('/api/user/<user_id>/messages')
def api_user_messages(user_id):
    messages = get_messages()
    user_messages = [m for m in messages if m['user_id'] == user_id]
    user_messages.sort(key=lambda x: x['message_id'], reverse=True)
    return jsonify(user_messages)

@app.route('/api/message/<message_id>')
def api_message(message_id):
    messages = get_messages()
    for m in messages:
        if m['message_id'] == message_id:
            users = get_users()
            for u in users:
                if u['user_id'] == m['user_id']:
                    m['nickname'] = u['nickname']
                    break
            return jsonify(m)
    return jsonify({'error': 'Message not found'}), 404

@app.route('/api/related/<message_id>')
def api_related(message_id):
    """关联广播：基于关键词匹配 + 加权随机"""
    users = get_users()
    messages = get_messages()
    
    # 找到当前广播
    current = None
    for m in messages:
        if m['message_id'] == message_id:
            current = m
            break
    if not current:
        return jsonify([])
    
    current_tag = parse_message_tag(current['message_tag'])['tag']
    if not current_tag:
        return jsonify([])
    
    # 筛选相关广播
    related = []
    for m in messages:
        if m['message_id'] == message_id:
            continue
        if not is_broadcast_visible(m):
            continue
        m_tag = parse_message_tag(m['message_tag'])['tag']
        if not m_tag:
            continue
        if is_related(current_tag, m_tag):
            related.append(m)
    
    if not related:
        return jsonify([])
    
    # 计算展示分并加权随机采样
    scores = [calculate_display_score(m, users) for m in related]
    selected = weighted_random_sample(related, scores, 3)
    
    # 补充昵称
    user_dict = {u['user_id']: u for u in users}
    for item in selected:
        user = user_dict.get(item['user_id'], {})
        item['nickname'] = user.get('nickname', item['user_id'].split('@')[0])
    
    return jsonify(selected)

@app.route('/api/search')
def api_search():
    """搜索：按关键词匹配 message_tag 或 user_id，返回按展示分排序的广播（最多20条）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    
    users = get_users()
    messages = get_messages()
    
    # 提取搜索词的关键词
    search_keywords = extract_keywords(q)
    if not search_keywords:
        return jsonify([])
    
    # 构建映射
    user_dict = {u['user_id']: u for u in users}
    
    # 筛选匹配的广播
    results = []
    for m in messages:
        if not is_broadcast_visible(m):
            continue
        
        # 检查 message_tag
        tag = parse_message_tag(m['message_tag'])['tag']
        tag_keywords = extract_keywords(tag)
        
        # 检查用户信息
        user = user_dict.get(m['user_id'])
        user_keywords = []
        if user:
            user_keywords.extend(extract_keywords(user.get('nickname', '')))
        user_keywords.extend(extract_keywords(m['user_id'].split('@')[0]))
        
        # 关键词匹配
        if set(search_keywords) & (set(tag_keywords) | set(user_keywords)):
            results.append(m)
    
    if not results:
        return jsonify([])
    
    # 计算展示分并排序
    for m in results:
        m['display_score'] = calculate_display_score(m, users)
    results.sort(key=lambda x: x['display_score'], reverse=True)
    
    # 取前20条并补充昵称
    for item in results[:20]:
        user = user_dict.get(item['user_id'])
        item['nickname'] = user.get('nickname', item['user_id'].split('@')[0]) if user else item['user_id'].split('@')[0]
    
    return jsonify(results[:20])


if __name__ == '__main__':
    ensure_excel_exists()
    print("Dot2Dot WebService 启动")
    print("API 地址: http://127.0.0.1:5000")
    print("  POST /api/login                      - 登录/注册")
    print("  GET  /api/users                      - 所有用户")
    print("  GET  /api/messages                   - 所有广播")
    print("  GET  /api/platform                   - 平台广场推荐")
    print("  GET  /api/user/<id>                  - 用户信息")
    print("  GET  /api/user/<id>/recommend        - 用户视图推荐")
    print("  GET  /api/user/<id>/messages         - 用户全部广播")
    print("  GET  /api/message/<id>               - 广播详情")
    print("  GET  /api/related/<id>               - 关联广播")
    print("  GET  /api/search?q=xxx               - 搜索")
    app.run(host='127.0.0.1', port=5000, debug=True)